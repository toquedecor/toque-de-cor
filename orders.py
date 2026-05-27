"""
Módulo de Pedidos — Toque de Cor Web

Responsabilidades:
  - Montar Excel por marca (Suvinil/Glasurit × Sherwin-Williams)
  - Envio de e-mail com pedido em HTML + anexos Excel
  - Exportação Suvinil: Cod Citel | SKU | Descrição | Quantidade
"""

import io
import os
import smtplib
import ssl
from datetime import datetime
from email.mime.application import MIMEApplication
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# Marcas que vão para o arquivo Suvinil/Glasurit
MARCAS_SUVINIL = {"SUVINIL", "GLASURIT", "GLASU", "GLASURITE"}
# Marcas que vão para o arquivo Sherwin-Williams
MARCAS_SW = {"SHERWIN", "SHERWIN-WILLIAMS", "SW", "LOXON", "METALATEX"}


def _classifica_marca(marca: str) -> str:
    """Retorna 'suvinil', 'sw' ou 'outros' com base na marca do item."""
    m = marca.upper().strip()
    for s in MARCAS_SUVINIL:
        if s in m:
            return "suvinil"
    for s in MARCAS_SW:
        if s in m:
            return "sw"
    return "outros"


# ── Exportação Excel ─────────────────────────────────────────────────────────
_COR_HEADER = "C0392B"   # vermelho Toque de Cor


def _escrever_planilha(ws, df: pd.DataFrame, pedido: dict | None = None) -> None:
    """Escreve cabeçalho de identificação + tabela de dados + ajusta largura das colunas."""
    # ── Bloco de identificação ───────────────────────────────────────────────
    data_start_row = 1
    if pedido:
        raw_dt = pedido.get("criado_em", "")
        try:
            dt = datetime.fromisoformat(str(raw_dt).replace("Z", "+00:00"))
            data_fmt = dt.strftime("%d/%m/%Y")
            hora_fmt = dt.strftime("%H:%M")
        except Exception:
            data_fmt = datetime.now().strftime("%d/%m/%Y")
            hora_fmt = datetime.now().strftime("%H:%M")

        meta = [
            ("Loja:",     pedido.get("loja", "")),
            ("Operador:", pedido.get("usuario", "")),
            ("UF:",       pedido.get("uf", "")),
            ("Data:",     data_fmt),
            ("Hora:",     hora_fmt),
        ]
        bold = Font(bold=True)
        for i, (label, value) in enumerate(meta, start=1):
            ws.cell(row=i, column=1, value=label).font = bold
            ws.cell(row=i, column=2, value=value)
        data_start_row = len(meta) + 2   # linha em branco de separação

    # ── Cabeçalho da tabela ──────────────────────────────────────────────────
    header_font  = Font(bold=True, color="FFFFFF")
    header_fill  = PatternFill(start_color=_COR_HEADER, end_color=_COR_HEADER, fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")

    for col_idx, col_name in enumerate(df.columns, start=1):
        cell = ws.cell(row=data_start_row, column=col_idx, value=col_name)
        cell.font    = header_font
        cell.fill    = header_fill
        cell.alignment = header_align
    ws.row_dimensions[data_start_row].height = 18

    # ── Dados ────────────────────────────────────────────────────────────────
    for row_idx, row in enumerate(df.itertuples(index=False), start=data_start_row + 1):
        for col_idx, value in enumerate(row, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    # ── Auto-size de colunas ─────────────────────────────────────────────────
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 4, 65)


def _exportar_excel_interativo(
    itens: list[dict],
    pedido: dict | None,
    marca_filtro: str,
    ws_title: str,
) -> bytes:
    """
    Gera Excel interativo com célula de Desconto Global editável (amarela) e
    fórmulas que recalculam Preço c/ Desc. e Total automaticamente.

    Estrutura:
      Linhas 1-5 : Loja, Operador, UF, Data, Hora
      Linha 6    : Desconto Global (%) | [célula amarela editável = desconto_pct do pedido]
      Linha 8    : Cabeçalho da tabela
      Linhas 9+  : Dados com fórmulas
      Última+2   : Total Geral
    """
    wb = Workbook()
    ws = wb.active
    ws.title = ws_title

    _RED  = "C0392B"
    _YELL = "FFFF00"
    _GRAY = "F0F0F0"

    bold         = Font(bold=True)
    hdr_font     = Font(bold=True, color="FFFFFF")
    hdr_fill     = PatternFill(start_color=_RED,  end_color=_RED,  fill_type="solid")
    hdr_align    = Alignment(horizontal="center", vertical="center")
    red_fill     = PatternFill(start_color=_RED,  end_color=_RED,  fill_type="solid")
    yellow_fill  = PatternFill(start_color=_YELL, end_color=_YELL, fill_type="solid")
    gray_fill    = PatternFill(start_color=_GRAY, end_color=_GRAY, fill_type="solid")
    money_fmt    = '"R$ "#,##0.00'
    pct_fmt      = "0.00"

    desconto_pct = float((pedido or {}).get("desconto_pct", 0))
    raw_dt = (pedido or {}).get("criado_em", "")
    try:
        dt = datetime.fromisoformat(str(raw_dt).replace("Z", "+00:00"))
        data_fmt = dt.strftime("%d/%m/%Y")
        hora_fmt = dt.strftime("%H:%M")
    except Exception:
        data_fmt = datetime.now().strftime("%d/%m/%Y")
        hora_fmt = datetime.now().strftime("%H:%M")

    # ── Linhas 1-5: metadados ────────────────────────────────────────────────
    meta = [
        ("Loja:",     (pedido or {}).get("loja", "")),
        ("Operador:", (pedido or {}).get("usuario", "")),
        ("UF:",       (pedido or {}).get("uf", "")),
        ("Data:",     data_fmt),
        ("Hora:",     hora_fmt),
    ]
    for i, (label, value) in enumerate(meta, start=1):
        ws.cell(row=i, column=1, value=label).font = bold
        ws.cell(row=i, column=2, value=value)

    # ── Linha 6: Desconto Global (célula amarela editável) ───────────────────
    lbl6 = ws.cell(row=6, column=1, value="Desconto Global (%):") 
    lbl6.font = Font(bold=True, color="FFFFFF")
    lbl6.fill = red_fill
    inp6 = ws.cell(row=6, column=2, value=desconto_pct)
    inp6.fill = yellow_fill
    inp6.number_format = pct_fmt

    # ── Linha 8: cabeçalho da tabela ────────────────────────────────────────
    incluir_marca = (marca_filtro == "outros")
    if incluir_marca:
        headers = [
            "Cod Citel", "SKU", "Marca", "Descrição", "Quantidade",
            "Preço Unit. (R$)", "Desconto (%)", "Preço c/ Desc. (R$)", "Total c/ Desc. (R$)",
        ]
        C_CITEL=1; C_SKU=2; C_MARCA=3; C_DESC=4; C_QTD=5
        C_PRECO=6; C_DPCT=7; C_PDISC=8; C_TOTAL=9
    else:
        headers = [
            "Cod Citel", "SKU", "Descrição", "Quantidade",
            "Preço Unit. (R$)", "Desconto (%)", "Preço c/ Desc. (R$)", "Total c/ Desc. (R$)",
        ]
        C_CITEL=1; C_SKU=2; C_DESC=3; C_QTD=4
        C_PRECO=5; C_DPCT=6; C_PDISC=7; C_TOTAL=8

    HDR = 8
    for ci, h in enumerate(headers, start=1):
        cell = ws.cell(row=HDR, column=ci, value=h)
        cell.font      = hdr_font
        cell.fill      = hdr_fill
        cell.alignment = hdr_align
    ws.row_dimensions[HDR].height = 18

    # ── Dados com fórmulas ───────────────────────────────────────────────────
    itens_marca = [it for it in itens if _classifica_marca(it.get("marca", "")) == marca_filtro]

    # Reverter desconto já aplicado para obter o preço base
    def _base_price(preco_unit: float) -> float:
        if desconto_pct <= 0 or desconto_pct >= 100:
            return preco_unit
        return round(preco_unit / (1 - desconto_pct / 100), 4)

    DS = 9  # data start row
    for offset, it in enumerate(itens_marca):
        r = DS + offset
        ws.cell(row=r, column=C_CITEL, value=it.get("cod_citel", ""))
        ws.cell(row=r, column=C_SKU,   value=it.get("cod_sku", ""))
        if incluir_marca:
            ws.cell(row=r, column=C_MARCA, value=it.get("marca", ""))
        ws.cell(row=r, column=C_DESC, value=it.get("descricao", ""))
        ws.cell(row=r, column=C_QTD,  value=int(it.get("qtd", 0)))

        bp = _base_price(float(it.get("preco_unit", 0)))
        pc = ws.cell(row=r, column=C_PRECO, value=bp)
        pc.number_format = money_fmt

        # Desconto (%) → referencia $B$6
        dc = ws.cell(row=r, column=C_DPCT, value="=$B$6")
        dc.number_format = pct_fmt

        # Preço c/ Desc. = Preço Unit. × (1 - $B$6/100)
        lp = get_column_letter(C_PRECO)
        pd_cell = ws.cell(row=r, column=C_PDISC, value=f"={lp}{r}*(1-$B$6/100)")
        pd_cell.number_format = money_fmt

        # Total c/ Desc. = Qtd × Preço c/ Desc.
        lq  = get_column_letter(C_QTD)
        lpd = get_column_letter(C_PDISC)
        tc  = ws.cell(row=r, column=C_TOTAL, value=f"={lq}{r}*{lpd}{r}")
        tc.number_format = money_fmt

    last = DS + len(itens_marca) - 1

    # ── Total Geral ──────────────────────────────────────────────────────────
    if itens_marca:
        tr = last + 2
        lbl_t = ws.cell(row=tr, column=C_PDISC, value="Total Geral:")
        lbl_t.font      = Font(bold=True)
        lbl_t.alignment = Alignment(horizontal="right")
        lt = get_column_letter(C_TOTAL)
        sum_t = ws.cell(row=tr, column=C_TOTAL, value=f"=SUM({lt}{DS}:{lt}{last})")
        sum_t.font          = Font(bold=True)
        sum_t.fill          = gray_fill
        sum_t.number_format = money_fmt

    # ── Auto-size de colunas ─────────────────────────────────────────────────
    for col in ws.columns:
        max_len = 0
        for cell in col:
            if cell.value is not None:
                try:
                    max_len = max(max_len, len(str(cell.value)))
                except Exception:
                    pass
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 65)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def exportar_excel_suvinil(
    itens: list[dict],
    pedido: dict | None = None,
    mostrar_precos: bool = False,
) -> bytes:
    """Gera .xlsx Suvinil/Glasurit.
    mostrar_precos=True  → formato interativo com célula de desconto e fórmulas.
    mostrar_precos=False → formato simples (Cod Citel, SKU, Descrição, Embalagem, Qtd).
    """
    if mostrar_precos:
        return _exportar_excel_interativo(itens, pedido, "suvinil", "Pedido Suvinil")
    rows = [
        {
            "Cod Citel":  it.get("cod_citel", ""),
            "SKU":        it.get("cod_sku", ""),
            "Descrição":  it.get("descricao", ""),
            "Embalagem":  it.get("embalagem", ""),
            "Quantidade": int(it.get("qtd", 0)),
        }
        for it in itens
        if _classifica_marca(it.get("marca", "")) == "suvinil"
    ]
    cols = ["Cod Citel", "SKU", "Descrição", "Embalagem", "Quantidade"]
    df = pd.DataFrame(rows, columns=cols) if rows else pd.DataFrame(columns=cols)
    wb = Workbook()
    ws = wb.active
    ws.title = "Pedido Suvinil"
    _escrever_planilha(ws, df, pedido)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def exportar_excel_sw(
    itens: list[dict],
    pedido: dict | None = None,
    mostrar_precos: bool = False,
) -> bytes:
    """Gera .xlsx Sherwin-Williams.
    mostrar_precos=True  → formato interativo com célula de desconto e fórmulas.
    mostrar_precos=False → formato simples.
    """
    if mostrar_precos:
        return _exportar_excel_interativo(itens, pedido, "sw", "Pedido SW")
    rows = [
        {
            "Cod Citel":  it.get("cod_citel", ""),
            "SKU":        it.get("cod_sku", ""),
            "Descrição":  it.get("descricao", ""),
            "Embalagem":  it.get("embalagem", ""),
            "Quantidade": int(it.get("qtd", 0)),
        }
        for it in itens
        if _classifica_marca(it.get("marca", "")) == "sw"
    ]
    cols = ["Cod Citel", "SKU", "Descrição", "Embalagem", "Quantidade"]
    df = pd.DataFrame(rows, columns=cols) if rows else pd.DataFrame(columns=cols)
    wb = Workbook()
    ws = wb.active
    ws.title = "Pedido SW"
    _escrever_planilha(ws, df, pedido)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def exportar_excel_outros(
    itens: list[dict],
    pedido: dict | None = None,
    mostrar_precos: bool = False,
) -> bytes:
    """Gera .xlsx para itens de outras marcas.
    mostrar_precos=True  → formato interativo com célula de desconto e fórmulas.
    mostrar_precos=False → formato simples.
    """
    if mostrar_precos:
        return _exportar_excel_interativo(itens, pedido, "outros", "Outras Marcas")
    rows = [
        {
            "Cod Citel":  it.get("cod_citel", ""),
            "SKU":        it.get("cod_sku", ""),
            "Marca":      it.get("marca", ""),
            "Descrição":  it.get("descricao", ""),
            "Embalagem":  it.get("embalagem", ""),
            "Quantidade": int(it.get("qtd", 0)),
        }
        for it in itens
        if _classifica_marca(it.get("marca", "")) == "outros"
    ]
    cols = ["Cod Citel", "SKU", "Marca", "Descrição", "Embalagem", "Quantidade"]
    df = pd.DataFrame(rows, columns=cols) if rows else pd.DataFrame(columns=cols)
    wb = Workbook()
    ws = wb.active
    ws.title = "Outras Marcas"
    _escrever_planilha(ws, df, pedido)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def exportar_excel_citel(itens: list[dict]) -> bytes:
    """
    Gera .xlsx no formato CITEL (sem cabeçalho, 8 colunas):
      Coluna B (índice 1): COD_CITEL
      Coluna F (índice 5): Quantidade
      Coluna H (índice 7): Preço com desconto já aplicado (preco_unit)
    Formato fixo esperado pelo sistema CITEL para importação — sem alterações.
    """
    rows = []
    for it in itens:
        row = [""] * 8
        row[1] = str(it.get("cod_citel", ""))
        row[5] = int(it.get("qtd", 0))
        row[7] = float(it.get("preco_unit", 0))
        rows.append(row)
    buf = io.BytesIO()
    pd.DataFrame(rows).to_excel(buf, index=False, header=False)
    buf.seek(0)
    return buf.read()


def exportar_excel_completo(
    itens: list[dict],
    mostrar_precos: bool = True,
    pedido: dict | None = None,
) -> bytes:
    """Exporta todos os itens do pedido em um único arquivo."""
    cols = ["Cod Citel", "SKU", "Marca", "Descrição", "Embalagem", "Quantidade"]
    if mostrar_precos:
        cols += ["Preço Unit. (R$)", "Total (R$)"]

    rows = []
    for it in itens:
        row = {
            "Cod Citel":  it.get("cod_citel", ""),
            "SKU":        it.get("cod_sku", ""),
            "Marca":      it.get("marca", ""),
            "Descrição":  it.get("descricao", ""),
            "Embalagem":  it.get("embalagem", ""),
            "Quantidade": int(it.get("qtd", 0)),
        }
        if mostrar_precos:
            row["Preço Unit. (R$)"] = float(it.get("preco_unit", 0))
            row["Total (R$)"]       = float(it.get("total", 0))
        rows.append(row)

    df = pd.DataFrame(rows, columns=cols)
    wb = Workbook()
    ws = wb.active
    ws.title = "Pedido"
    _escrever_planilha(ws, df, pedido)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


# ── E-mail ───────────────────────────────────────────────────────────────────
def _config_smtp() -> dict:
    """Lê configurações SMTP/SendGrid do Supabase ou variáveis de ambiente."""
    cfg = {
        "host":             os.environ.get("SMTP_HOST", "smtp.gmail.com"),
        "port":             int(os.environ.get("SMTP_PORT", "587")),
        "usuario":          os.environ.get("SMTP_USUARIO", ""),
        "senha":            os.environ.get("SMTP_SENHA", ""),
        "remetente":        os.environ.get("SMTP_REMETENTE", ""),
        "destinatarios":    os.environ.get("SMTP_DESTINATARIOS", ""),
        "sendgrid_api_key": os.environ.get("SENDGRID_API_KEY", ""),
    }
    # Sobrescreve com valores do banco se disponíveis
    try:
        from db_supabase import get_config
        for k in ("host", "port", "usuario", "senha", "remetente", "destinatarios"):
            val = get_config(f"smtp_{k}")
            if val:
                cfg[k] = int(val) if k == "port" else val
        # SendGrid API key (salva como 'smtp_sendgrid_api_key' no banco)
        sg = get_config("smtp_sendgrid_api_key")
        if sg:
            cfg["sendgrid_api_key"] = sg
    except Exception:
        pass
    return cfg


def _enviar_via_sendgrid(
    api_key: str,
    from_email: str,
    destinos: list[str],
    subject: str,
    html: str,
    anexos: list[tuple[bytes, str]],
) -> None:
    """
    Envia e-mail via SendGrid HTTP API (porta 443 — funciona no HuggingFace).
    Lança Exception em caso de falha.
    """
    import base64
    import json
    import urllib.request

    payload: dict = {
        "personalizations": [{"to": [{"email": d} for d in destinos]}],
        "from":    {"email": from_email},
        "subject": subject,
        "content": [{"type": "text/html", "value": html}],
    }
    if anexos:
        payload["attachments"] = [
            {
                "content":     base64.b64encode(data).decode(),
                "filename":    filename,
                "type":        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                "disposition": "attachment",
            }
            for data, filename in anexos
        ]

    req = urllib.request.Request(
        "https://api.sendgrid.com/v3/mail/send",
        data=json.dumps(payload).encode("utf-8"),
        headers={
            "Authorization": f"Bearer {api_key}",
            "Content-Type":  "application/json",
        },
        method="POST",
    )
    with urllib.request.urlopen(req, timeout=15) as resp:
        status = resp.status
    if status not in (200, 202):
        raise Exception(f"SendGrid retornou HTTP {status}")


def _html_pedido(pedido: dict, itens: list[dict], mostrar_precos: bool) -> str:
    """Gera corpo HTML do e-mail de pedido."""
    data = pedido.get("criado_em", datetime.now().strftime("%d/%m/%Y %H:%M"))
    num  = pedido.get("numero", 0)
    uf   = pedido.get("uf", "")
    loja = pedido.get("loja", "")
    usr  = pedido.get("usuario", "")

    linhas_html = ""
    total_geral = 0.0
    for it in itens:
        total = float(it.get("total", 0))
        total_geral += total
        preco_col = (
            f"<td>R$ {it.get('preco_unit', 0):.2f}</td>"
            f"<td><b>R$ {total:.2f}</b></td>"
            if mostrar_precos else ""
        )
        linhas_html += (
            f"<tr>"
            f"<td>{it.get('cod_citel','')}</td>"
            f"<td>{it.get('cod_sku','')}</td>"
            f"<td>{it.get('marca','')}</td>"
            f"<td>{it.get('descricao','')}</td>"
            f"<td>{it.get('embalagem','')}</td>"
            f"<td>{it.get('qtd',0)}</td>"
            f"{preco_col}"
            f"</tr>"
        )

    preco_header = "<th>Preço Unit.</th><th>Total</th>" if mostrar_precos else ""
    total_row = (
        f"<tr style='font-weight:bold;background:#f0f0f0'>"
        f"<td colspan='5' style='text-align:right'>TOTAL GERAL</td>"
        f"<td></td><td></td><td>R$ {total_geral:,.2f}</td></tr>"
        if mostrar_precos else ""
    )

    return f"""
    <html><body style="font-family:Arial,sans-serif;color:#333">
    <h2 style="color:#e63946">🎨 Toque de Cor — Pedido #{num:04d}</h2>
    <table style="margin-bottom:12px">
      <tr><td><b>Loja:</b></td><td>{loja}</td></tr>
      <tr><td><b>Operador:</b></td><td>{usr}</td></tr>
      <tr><td><b>UF:</b></td><td>{uf}</td></tr>
      <tr><td><b>Data:</b></td><td>{data}</td></tr>
    </table>
    <table border="1" cellpadding="6" cellspacing="0"
           style="border-collapse:collapse;width:100%;font-size:13px">
      <thead style="background:#e63946;color:#fff">
        <tr>
          <th>Cod Citel</th><th>SKU</th><th>Marca</th>
          <th>Descrição</th><th>Embalagem</th><th>Qtd</th>
          {preco_header}
        </tr>
      </thead>
      <tbody>{linhas_html}</tbody>
      {total_row}
    </table>
    <p style="color:#888;font-size:12px;margin-top:16px">
      Mensagem gerada automaticamente — Toque de Cor Sistema de Pedidos
    </p>
    </body></html>
    """


def enviar_email_pedido(
    pedido: dict,
    itens: list[dict],
    mostrar_precos: bool = True,
    destinatarios_extra: list[str] | None = None,
) -> tuple[bool, str]:
    """
    Envia o pedido por e-mail com anexos Excel separados por marca.
    Usa SendGrid HTTP API se a chave estiver configurada (funciona no HuggingFace).
    Fallback para SMTP clássico caso contrário.
    Retorna (sucesso, mensagem).
    """
    cfg    = _config_smtp()
    sg_key = cfg.get("sendgrid_api_key", "")

    # Precisa de SendGrid OU credenciais SMTP
    if not sg_key and (not cfg["usuario"] or not cfg["senha"]):
        return False, (
            "E-mail não configurado. "
            "Acesse Painel Admin → Configurações e informe a SendGrid API Key."
        )

    destinos_str = cfg.get("destinatarios", "")
    destinos = [d.strip() for d in destinos_str.split(",") if d.strip()]
    if destinatarios_extra:
        destinos += [d for d in destinatarios_extra if d.strip()]
    if not destinos:
        return False, "Nenhum destinatário configurado."

    num        = pedido.get("numero", 0)
    loja       = pedido.get("loja", "")
    data       = datetime.now().strftime("%d-%m-%Y")
    subject    = f"Pedido #{num:04d} — {loja} — {data}"
    from_email = cfg.get("remetente") or cfg["usuario"]
    html_body  = _html_pedido(pedido, itens, mostrar_precos)

    # Monta anexos Excel separados por marca + planilhas CITEL
    _itens_suv    = [it for it in itens if _classifica_marca(it.get("marca", "")) == "suvinil"]
    _itens_sw     = [it for it in itens if _classifica_marca(it.get("marca", "")) == "sw"]
    _itens_outros = [it for it in itens if _classifica_marca(it.get("marca", "")) == "outros"]

    anexos: list[tuple[bytes, str]] = []
    if _itens_suv:
        anexos.append((
            exportar_excel_suvinil(itens, pedido=pedido, mostrar_precos=mostrar_precos),
            f"Pedido_Suvinil_{loja}_{data}.xlsx",
        ))
        anexos.append((
            exportar_excel_citel(_itens_suv),
            f"Importacao_Citel_Suvinil_{loja}_{data}.xlsx",
        ))
    if _itens_sw:
        anexos.append((
            exportar_excel_sw(itens, pedido=pedido, mostrar_precos=mostrar_precos),
            f"Pedido_SW_{loja}_{data}.xlsx",
        ))
        anexos.append((
            exportar_excel_citel(_itens_sw),
            f"Importacao_Citel_SW_{loja}_{data}.xlsx",
        ))
    if _itens_outros:
        anexos.append((
            exportar_excel_outros(itens, pedido=pedido, mostrar_precos=mostrar_precos),
            f"Pedido_Outros_{loja}_{data}.xlsx",
        ))
        anexos.append((
            exportar_excel_citel(_itens_outros),
            f"Importacao_Citel_Outros_{loja}_{data}.xlsx",
        ))

    # ── SendGrid (prioridade — funciona no HuggingFace) ──────────────────────
    if sg_key:
        try:
            _enviar_via_sendgrid(sg_key, from_email, destinos, subject, html_body, anexos)
            return True, f"E-mail enviado para: {', '.join(destinos)}"
        except Exception as e:
            return False, f"Erro SendGrid: {e}"

    # ── SMTP fallback (ambientes sem bloqueio de porta) ───────────────────────
    msg = MIMEMultipart("mixed")
    msg["Subject"] = subject
    msg["From"]    = from_email
    msg["To"]      = ", ".join(destinos)
    msg.attach(MIMEText(html_body, "html", "utf-8"))
    for xls_bytes, filename in anexos:
        att = MIMEApplication(xls_bytes, _subtype="xlsx")
        att.add_header("Content-Disposition", "attachment", filename=filename)
        msg.attach(att)
    try:
        ctx = ssl.create_default_context()
        with smtplib.SMTP(cfg["host"], cfg["port"], timeout=15) as s:
            s.starttls(context=ctx)
            s.login(cfg["usuario"], cfg["senha"])
            s.sendmail(from_email, destinos, msg.as_bytes())
        return True, f"E-mail enviado para: {', '.join(destinos)}"
    except smtplib.SMTPAuthenticationError:
        return False, "Falha de autenticação SMTP. Verifique usuário/senha."
    except Exception as e:
        return False, f"Erro ao enviar e-mail: {e}"
